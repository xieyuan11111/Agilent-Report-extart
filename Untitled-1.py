import os
import pandas as pd
from openpyxl import Workbook
import tempfile
import shutil


def find_excel_files(directory):
    excel_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file == "Report01.xls":
                excel_files.append((os.path.join(root, file), os.path.basename(root)))
    return excel_files


def extract_area_data(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='Peak')
        return df['Area'].tolist()
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return []


def write_to_xlsx(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Area Data"

    for col, (folder_name, values) in enumerate(data.items(), start=1):
        ws.cell(row=1, column=col, value=folder_name)
        for row, value in enumerate(values, start=2):
            ws.cell(row=row, column=col, value=value)

    # 使用临时文件来保存
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_name = tmp.name
        wb.save(tmp_name)

    # 尝试移动临时文件到目标位置
    try:
        shutil.move(tmp_name, output_file)
        print(f"Data has been extracted and written to {output_file}")
    except PermissionError:
        print(f"无法保存到 {output_file}。文件可能正在使用中。")
        print(f"数据已保存到临时文件: {tmp_name}")
    except Exception as e:
        print(f"保存文件时发生错误: {str(e)}")
        print(f"数据已保存到临时文件: {tmp_name}")


def main():
    current_directory = os.getcwd()
    excel_files = find_excel_files(current_directory)

    all_area_data = {}
    for file_path, folder_name in excel_files:
        area_data = extract_area_data(file_path)
        all_area_data[folder_name] = area_data

    output_file = "extracted_area_data.xlsx"
    write_to_xlsx(all_area_data, output_file)


if __name__ == "__main__":
    main()